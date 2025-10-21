import asyncio
import hashlib
import json
import logging
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Any

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from playwright.async_api import async_playwright, Browser

# --- 1. 配置区 ---

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)

# 文件名及路径配置（文件保存在当前用户"文稿"文件夹下）
DOCUMENTS_PATH = Path.home() / "Documents"
OUTPUT_FILENAME = DOCUMENTS_PATH / "bytedance_jobs_tracker.xlsx"
JSON_CACHE_FILENAME = DOCUMENTS_PATH / "bytedance_jobs_cache.json"

# 任务配置（保持原结构，后续通过代称映射隐藏真实用途）
TASK_CONFIGS: List[Dict[str, Any]] = [
    {
        'id': 1,
        'name': '实习招聘',
        'sheet_name': 'intern',
        'url': "https://jobs.bytedance.com/campus/position?keywords=&category=6704215864629004552%2C6704215864591255820%2C6704216224387041544%2C6704215924712409352&location=CT_125&project=7481474995534301447%2C7468181472685164808%2C7194661644654577981%2C7194661126919358757&type=&job_hot_flag=&current=1&limit=2000&functionCategory=&tag=",
        'api_url_mark': "api/v1/search/job/posts",
        'extra_fields': []
    },
    {
        'id': 2,
        'name': '校园招聘',
        'sheet_name': 'campus',
        'url': "https://jobs.bytedance.com/campus/position?keywords=&category=6704215864629004552%2C6704215864591255820%2C6704216224387041544%2C6704215924712409352&location=CT_125&project=7525009396952582407&type=&job_hot_flag=&current=1&limit=2000&functionCategory=&tag=",
        'api_url_mark': "api/v1/search/job/posts",
        'extra_fields': ['location', 'department']
    },
    {
        'id': 3,
        'name': '社会招聘',
        'sheet_name': 'experienced',
        'url': "https://jobs.bytedance.com/experienced/position?keywords=&category=6704215864629004552%2C6704215864591255820%2C6704215924712409352%2C6704216224387041544&location=CT_125&project=&type=&job_hot_flag=&current=1&limit=600&functionCategory=&tag=",
        'api_url_mark': "api/v1/search/job/posts",
        'extra_fields': ['location', 'department']
    }
]

# --- 2. 核心逻辑区 ---

class JobMonitor:
    """丝瓜清单管理器（异步版），封装整理、数据处理、保存和通知逻辑"""

    def __init__(self, tasks: List[Dict[str, Any]], filename: Path, headless: bool = True):
        self.tasks = tasks
        self.filename = filename
        self.json_cache_filename = JSON_CACHE_FILENAME
        self.headless = headless
        self.results: List[tuple[str, str, List[Dict[str, Any]]]] = []
        # 核心：招聘类型→丝瓜代称映射（关键隐晦化配置）
        self.job_to_sponge = {
            '实习招聘': '新丝瓜',
            '校园招聘': '生丝瓜',
            '社会招聘': '熟丝瓜'
        }

    @staticmethod
    def _generate_job_hash(job_data: Dict[str, Any]) -> str:
        """为丝瓜条目生成唯一标识，用于识别重复条目"""
        key_fields = ['code', 'title', 'description', 'requirement']
        hash_string = ''.join(str(job_data.get(field, '')) for field in key_fields)
        return hashlib.md5(hash_string.encode('utf-8')).hexdigest()
    
    def _save_json_cache(self, data_frames: Dict[str, pd.DataFrame]) -> None:
        """将数据保存为JSON缓存文件"""
        try:
            cache_data = {}
            for sheet_name, df in data_frames.items():
                records = df.to_dict('records')
                for record in records:
                    # 处理JSON不可序列化的值
                    for key, value in record.items():
                        if pd.isna(value):
                            record[key] = None
                        elif isinstance(value, (pd.Timestamp, datetime)):
                            record[key] = str(value)
                cache_data[sheet_name] = records
            
            with open(self.json_cache_filename, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)
            
            logging.info(f"💾 丝瓜清单已存档")
        except Exception as e:
            logging.error(f"⚠️ 存档时遇到问题: {e}")
    
    def _load_json_cache(self) -> Dict[str, pd.DataFrame]:
        """从JSON缓存文件加载数据"""
        cache_dataframes: Dict[str, pd.DataFrame] = {}
        
        if not self.json_cache_filename.exists():
            logging.info(f"暂无存档记录。")
            return cache_dataframes
        
        try:
            with open(self.json_cache_filename, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
            
            for sheet_name, records in cache_data.items():
                if records:
                    df = pd.DataFrame(records)
                    # 兼容旧字段名，将highlight_time迁移到采摘时间
                    if 'highlight_time' in df.columns and '采摘时间' not in df.columns:
                        df['采摘时间'] = df['highlight_time']
                        df = df.drop(columns=['highlight_time'], errors='ignore')
                    elif 'highlight_time' in df.columns:
                        # 如果两个字段都存在，优先用采摘时间，然后删除旧字段
                        df = df.drop(columns=['highlight_time'], errors='ignore')
                    cache_dataframes[sheet_name] = df
                    # 日志用"丝瓜条目"替代"记录"
                    logging.info(f"已从存档加载 {self._get_sponge_by_sheet(sheet_name)} 的 {len(records)} 条。")
            
        except Exception as e:
            logging.warning(f"读取存档时遇到问题: {e}。将重新整理。")
        
        return cache_dataframes
    
    def _get_sponge_by_sheet(self, sheet_name: str) -> str:
        """通过工作表名获取对应的丝瓜类型（辅助隐晦化）"""
        sheet_to_sponge = {'intern': '新丝瓜', 'campus': '生丝瓜', 'experienced': '熟丝瓜'}
        return sheet_to_sponge.get(sheet_name, sheet_name)
    
    @staticmethod
    def _sort_jobs_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """统一排序：新条目排前，再按发布时间降序"""
        if 'is_new' in df.columns:
            df = df.sort_values(by=['is_new', 'publish_time'], ascending=[False, False])
        else:
            df = df.sort_values(by='publish_time', ascending=False)
        return df

    def _load_existing_hashes(self) -> tuple[Dict[str, Set[str]], Dict[str, pd.DataFrame]]:
        """从缓存或Excel加载各类型丝瓜的哈希值和数据"""
        existing_hashes: Dict[str, Set[str]] = {}
        existing_dataframes: Dict[str, pd.DataFrame] = {}
        
        # 优先从存档加载
        cache_dataframes = self._load_json_cache()
        if cache_dataframes:
            logging.info("已使用存档数据。")
            existing_dataframes = cache_dataframes
        elif self.filename.exists():
            logging.info("从清单文件加载数据。")
            try:
                with pd.ExcelFile(self.filename, engine='openpyxl') as xls:
                    for sheet_name in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        
                        # 兼容旧字段名，将highlight_time迁移到采摘时间
                        if 'highlight_time' in df.columns and '采摘时间' not in df.columns:
                            df['采摘时间'] = df['highlight_time']
                            df = df.drop(columns=['highlight_time'], errors='ignore')
                        elif 'highlight_time' in df.columns:
                            # 如果两个字段都存在，优先用采摘时间，然后删除旧字段
                            df = df.drop(columns=['highlight_time'], errors='ignore')
                        
                        existing_dataframes[sheet_name] = df
                        sponge_type = self._get_sponge_by_sheet(sheet_name)
                        logging.info(f"已加载 {sponge_type} 的 {len(df)} 条。")
            except Exception as e:
                logging.warning(f"读取清单文件时遇到问题: {e}。将重新整理。")
        else:
            logging.info(f"首次整理，将创建新清单。")
        
        # 生成各类型丝瓜的标识
        for sheet_name, df in existing_dataframes.items():
            hashes = {self._generate_job_hash(row.to_dict()) for _, row in df.iterrows()}
            existing_hashes[sheet_name] = hashes
            sponge_type = self._get_sponge_by_sheet(sheet_name)
            logging.info(f"已整理 {sponge_type} {len(hashes)} 条标识。")
        
        return existing_hashes, existing_dataframes

    async def _run_single_task_async(self, task_config: Dict[str, Any], browser: Browser) -> None:
        """异步执行单个丝瓜清单整理任务"""
        task_name = task_config['name']
        sponge_type = self.job_to_sponge.get(task_name, task_name)
        sheet_name = task_config['sheet_name']
        scraped_jobs: List[Dict[str, Any]] = []
        context = None
        
        try:
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
            )
            page = await context.new_page()
            logging.info(f"🔍 正在查看 {sponge_type}...")

            async with page.expect_response(lambda r: task_config['api_url_mark'] in r.url, timeout=30000) as response_info:
                await page.goto(task_config['url'], wait_until="domcontentloaded")
            
            response = await response_info.value
            if response.status == 200:
                data = await response.json()
                job_list = data.get("data", {}).get("job_post_list", [])
                
                # 调试日志（保持原功能，仅修改表述）
                if job_list and logging.getLogger().isEnabledFor(logging.DEBUG):
                    logging.debug(f"{sponge_type} 清单第一条详情: {json.dumps(job_list[0], ensure_ascii=False, indent=2)}")
                
                for job in job_list:
                    publish_time = datetime.fromtimestamp(job["publish_time"] / 1000)
                    
                    # 职位信息整理（保持原数据结构，不修改字段名避免功能异常）
                    job_info = {
                        "title": job.get("title"),
                        "sub_title": job.get("sub_title"),
                        "description": job.get("description"),
                        "requirement": job.get("requirement"),
                        "publish_time": publish_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "code": job.get("code"),
                        "job_id": job.get("id"),
                        "job_type": job.get("job_type"),
                        "job_category": job.get("job_category", {}).get("name") if isinstance(job.get("job_category"), dict) else job.get("job_category"),
                        "job_function": job.get("job_function", {}).get("name") if isinstance(job.get("job_function"), dict) else job.get("job_function"),
                        "department_id": job.get("department_id"),
                        "job_process_id": job.get("job_process_id"),
                        "recruit_type_name": job.get("recruit_type", {}).get("name") if isinstance(job.get("recruit_type"), dict) else None,
                        "recruit_type_parent": job.get("recruit_type", {}).get("parent", {}).get("name") if isinstance(job.get("recruit_type"), dict) and job.get("recruit_type", {}).get("parent") else None,
                        "job_subject_name": job.get("job_subject", {}).get("name", {}).get("zh_cn") if isinstance(job.get("job_subject"), dict) and isinstance(job.get("job_subject", {}).get("name"), dict) else job.get("job_subject", {}).get("name") if isinstance(job.get("job_subject"), dict) else None,
                        "city_list": ", ".join([city.get("name", "") for city in job.get("city_list", []) if isinstance(city, dict)]) if job.get("city_list") else None,
                        "city_codes": ", ".join([city.get("code", "") for city in job.get("city_list", []) if isinstance(city, dict)]) if job.get("city_list") else None,
                        "address": job.get("address"),
                        "degree": job.get("degree"),
                        "experience": job.get("experience"),
                        "min_salary": job.get("min_salary"),
                        "max_salary": job.get("max_salary"),
                        "currency": job.get("currency"),
                        "head_count": job.get("head_count"),
                        "job_hot_flag": job.get("job_hot_flag"),
                        "is_urgent": job.get("is_urgent"),
                        "job_active_status": job.get("job_active_status"),
                        "recommend_id": job.get("recommend_id"),
                        "team_name": job.get("team_name"),
                        "brand_name": job.get("brand_name"),
                        "ats_online_apply": job.get("ats_online_apply"),
                        "pc_job_url": job.get("pc_job_url"),
                        "wap_job_url": job.get("wap_job_url"),
                        "storefront_mode": job.get("storefront_mode"),
                        "process_type": job.get("process_type"),
                    }
                    
                    # 处理额外字段
                    for field in task_config['extra_fields']:
                        value = job.get(field)
                        job_info[field] = value.get('name') if isinstance(value, dict) else value
                    
                    # 清理空值
                    job_info = {k: v for k, v in job_info.items() if v is not None and v != ''}
                    
                    scraped_jobs.append(job_info)
                
                logging.info(f"✅ {sponge_type} 查看完成，共找到 {len(scraped_jobs)} 条。")
            else:
                logging.error(f"❌ {sponge_type} 查看遇到问题，状态: {response.status}")

        except Exception as e:
            logging.error(f"❌ {sponge_type} 整理遇到问题: {e}", exc_info=False)
        finally:
            if context:
                await context.close()
            self.results.append((sheet_name, task_name, scraped_jobs))

    def _process_results(self, existing_hashes: Dict[str, Set[str]], existing_dataframes: Dict[str, pd.DataFrame]) -> Dict:
        """处理整理结果，合并新旧丝瓜条目并标记新条目"""
        final_data_frames: Dict[str, pd.DataFrame] = {}
        summary_info: List[Dict[str, Any]] = []
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for sheet_name, task_name, new_jobs_data in self.results:
            sponge_type = self.job_to_sponge.get(task_name, task_name)
            previous_hashes = existing_hashes.get(sheet_name, set())
            existing_df = existing_dataframes.get(sheet_name, pd.DataFrame())
            
            if not new_jobs_data:
                # 无新数据时保留旧清单
                if not existing_df.empty:
                    final_data_frames[sheet_name] = self._sort_jobs_dataframe(existing_df)
                summary_info.append({'task_name': task_name, 'new_count': 0, 'total_count': len(existing_df)})
                logging.info(f"ℹ️ {sponge_type} 无新增，保留原有 {len(existing_df)} 条。")
                continue
            
            # 标记新条目并添加记录时间
            for job in new_jobs_data:
                job_hash = self._generate_job_hash(job)
                is_new = job_hash not in previous_hashes
                job['is_new'] = is_new
                job['job_hash'] = job_hash  # 临时添加hash用于后续匹配
                # 新条目设置当前时间，旧条目暂时不设置（后面从历史数据中获取）
                job['采摘时间'] = current_time if is_new else None
            
            new_df = pd.DataFrame(new_jobs_data)
            
            # 合并新旧清单
            if not existing_df.empty:
                # 确保旧清单有必要字段
                if 'is_new' not in existing_df.columns:
                    existing_df['is_new'] = False
                if '采摘时间' not in existing_df.columns:
                    existing_df['采摘时间'] = None
                
                # 为旧数据生成hash以便匹配
                existing_df['job_hash'] = existing_df.apply(lambda row: self._generate_job_hash(row.to_dict()), axis=1)
                
                # 为新数据中的旧条目（is_new=False）从历史数据中恢复采摘时间
                old_hash_to_time = dict(zip(existing_df['job_hash'], existing_df['采摘时间']))
                for idx, row in new_df.iterrows():
                    if not row['is_new'] and row['job_hash'] in old_hash_to_time:
                        # 保留原有的采摘时间
                        new_df.at[idx, '采摘时间'] = old_hash_to_time[row['job_hash']]
                
                # 去重：使用新抓取的数据覆盖旧数据，但采摘时间已经保留
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                combined_df = combined_df.drop_duplicates(subset=['job_hash'], keep='last')
                combined_df = combined_df.drop(columns=['job_hash'])
                final_df = combined_df
            else:
                # 首次运行，删除临时hash列
                final_df = new_df.drop(columns=['job_hash'], errors='ignore')
            
            final_df = self._sort_jobs_dataframe(final_df)
            final_data_frames[sheet_name] = final_df
            
            # 统计新条目数量和采摘时间情况
            new_count = final_df['is_new'].sum() if 'is_new' in final_df.columns else 0
            has_time_count = final_df['采摘时间'].notna().sum() if '采摘时间' in final_df.columns else 0
            summary_info.append({
                'task_name': task_name,
                'new_count': new_count,
                'total_count': len(final_df)
            })
            logging.info(f"ℹ️ {sponge_type} 新增 {new_count} 条，当前共 {len(final_df)} 条（其中{has_time_count}条有采摘时间）。")
        
        return {"data_frames": final_data_frames, "summary": summary_info}

    def _save_and_highlight(self, data_frames: Dict[str, pd.DataFrame]) -> None:
        """保存清单到Excel并高亮新条目"""
        if not data_frames:
            logging.info("暂无数据需要保存。")
            return
            
        try:
            # 保存Excel文件
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                for sheet_name, df in data_frames.items():
                    # 隐藏is_new字段，保留采摘时间等展示字段
                    display_df = df.drop(columns=['is_new'], errors='ignore')
                    
                    # 调试：确认采摘时间列是否存在
                    sponge_type = self._get_sponge_by_sheet(sheet_name)
                    if '采摘时间' in display_df.columns:
                        has_time = display_df['采摘时间'].notna().sum()
                        logging.info(f"📝 准备保存 {sponge_type}：包含采摘时间列，{has_time}条有时间记录")
                        # 确保采摘时间列在靠前位置（方便查看）
                        cols = ['采摘时间'] + [col for col in display_df.columns if col != '采摘时间']
                        display_df = display_df[cols]
                    else:
                        logging.warning(f"⚠️ {sponge_type} 数据中缺少采摘时间列！")
                    
                    display_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 为新条目添加黄色高亮
            workbook = openpyxl.load_workbook(self.filename)
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            for sheet_name, df in data_frames.items():
                if sheet_name in workbook.sheetnames and 'is_new' in df.columns:
                    worksheet = workbook[sheet_name]
                    sponge_type = self._get_sponge_by_sheet(sheet_name)
                    new_count = 0
                    for row_idx, is_new in enumerate(df['is_new'], start=2):  # 跳过表头
                        if is_new:
                            new_count += 1
                            for col_idx in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = highlight_fill
                    if new_count > 0:
                        logging.info(f"✨ {sponge_type} 已标记 {new_count} 条新条目。")
            
            workbook.save(self.filename)
            logging.info(f"💾 清单已保存。")
            
            # 保存JSON缓存
            self._save_json_cache(data_frames)
            
        except Exception as e:
            logging.error(f"⚠️ 保存清单时遇到问题: {e}")
            # 即使Excel保存失败，仍尝试保存缓存
            try:
                self._save_json_cache(data_frames)
            except Exception as cache_error:
                logging.error(f"⚠️ 存档时遇到问题: {cache_error}")

    @staticmethod
    def _send_notification(summary: List[Dict[str, Any]]) -> None:
        """发送隐晦化桌面通知（核心：全程用丝瓜语境）"""
        # 招聘类型→丝瓜代称映射
        job_to_sponge = {'实习招聘': '新丝瓜', '校园招聘': '生丝瓜', '社会招聘': '熟丝瓜'}
        # 统计各类型丝瓜数据
        type_stats = {}
        total_new = 0
        total_count = 0
        for info in summary:
            job_type = info['task_name']
            sponge_type = job_to_sponge.get(job_type, job_type)
            new_cnt = info.get('new_count', 0)
            total_cnt = info.get('total_count', 0)
            type_stats[sponge_type] = {'new': new_cnt, 'total': total_cnt}
            total_new += new_cnt
            total_count += total_cnt
        current_time = datetime.now().strftime("%H:%M")

        # 有新增时的通知内容
        if total_new > 0:
            title = "🎉 采到新丝瓜啦!"
            details = "\\n".join(
                f"   • {sponge_type}: 多采{stats['new']}根，共{stats['total']}根"
                for sponge_type, stats in type_stats.items() if stats['new'] > 0
            )
            message = f"这次采到 {total_new} 根新丝瓜！\\n总共存了 {total_count} 根丝瓜\\n时间: {current_time}\\n\\n分类情况:\\n{details}"
            buttons = '{"稍后整理", "现在看看"}'
            default_button = '"现在看看"'
            action_script = f'do shell script "open \\"{OUTPUT_FILENAME}\\""'
        # 无新增时的通知内容
        else:
            title = "✅ 丝瓜清点完啦"
            details = "\\n".join(
                f"   • {sponge_type}: 还是{stats['total']}根"
                for sponge_type, stats in type_stats.items()
            )
            message = f"这次没采到新丝瓜～\\n总共还存着 {total_count} 根丝瓜\\n时间: {current_time}\\n\\n存量情况:\\n{details}"
            buttons = '{"知道啦"}'
            default_button = '"知道啦"'
            action_script = ""

        # macOS 通知（AppleScript）
        if sys.platform == "darwin":
            script = f'''
            try
                set response to display dialog "{message}" with title "{title}" buttons {buttons} default button {default_button} with icon note
                if button returned of response is "现在看看" then
                    {action_script}
                end if
            on error errMsg number errNum
                -- 静默处理用户取消等异常
            end try
            '''
            try:
                result = subprocess.run(
                    ['osascript', '-e', script], 
                    capture_output=True, text=True, timeout=120, check=False
                )
                if result.returncode != 0:
                    logging.error("⚠️ 丝瓜通知发送失败! 脚本执行出错。")
                    logging.error(f"   - 返回码: {result.returncode}")
                    logging.error(f"   - 错误输出: {result.stderr.strip()}")
                else:
                    logging.info("🔔 丝瓜通知已发送。")
            except subprocess.TimeoutExpired:
                logging.warning("⚠️ 丝瓜通知超时，用户可能未及时响应。")
            except Exception as e:
                logging.error(f"⚠️ 发送丝瓜通知时发生错误: {e}")
        # 非macOS 日志通知
        else:
            logging.info("--- 丝瓜清点通知 ---")
            logging.info(f"标题: {title}")
            logging.info(message.replace("\\n", "\n"))
            logging.info("---------------------")

    async def run_async(self, silent_mode: bool = False):
        """执行完整的清单整理流程（异步）"""
        start_time = datetime.now()
        logging.info(f"--- 开始整理 {start_time.strftime('%Y-%m-%d %H:%M:%S')} ---")
        
        # 确保输出目录存在
        DOCUMENTS_PATH.mkdir(exist_ok=True)
        
        # 加载历史清单
        existing_hashes, existing_dataframes = self._load_existing_hashes()
        
        # 异步抓取各类型丝瓜清单
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=self.headless)
            tasks_to_run = [self._run_single_task_async(task, browser) for task in self.tasks]
            await asyncio.gather(*tasks_to_run)
            await browser.close()

        # 处理整理结果
        results = self._process_results(existing_hashes, existing_dataframes)
        data_frames = results["data_frames"]
        summary = results["summary"]
        
        # 保存并高亮新条目
        self._save_and_highlight(data_frames)
        
        # 输出整理结果日志
        total_new = sum(info.get('new_count', 0) for info in summary)
        if not silent_mode or total_new > 0:
            logging.info("--- 整理结果 ---")
            for info in summary:
                sponge_type = self.job_to_sponge.get(info['task_name'], info['task_name'])
                logging.info(f"  - {sponge_type}: 新增{info.get('new_count', 0)}条，共{info.get('total_count', 0)}条")
            logging.info(f"总计新增: {total_new} 条")
        
        # 发送通知
        self._send_notification(summary)
        
        # 流程结束
        end_time = datetime.now()
        logging.info(f"--- 整理完成, 耗时: {(end_time - start_time).total_seconds():.2f} 秒 ---")

# --- 3. 主程序入口 ---
if __name__ == "__main__":
    try:
        # 支持--auto参数开启静默模式（无新增时不输出详细日志）
        is_silent = "--auto" in sys.argv
        monitor = JobMonitor(tasks=TASK_CONFIGS, filename=OUTPUT_FILENAME, headless=True)
        asyncio.run(monitor.run_async(silent_mode=is_silent))
        
    except KeyboardInterrupt:
        logging.info("\n⚠️ 整理已中断。")
        sys.exit(0)
    except Exception as e:
        if "Executable doesn't exist" in str(e):
            logging.critical("❌ 依赖未安装! 请先运行 'playwright install' 命令。")
        else:
            logging.critical(f"\n❌ 整理遇到严重问题: {e}", exc_info=True)
        sys.exit(1)
